"""
Daily refresh pipeline for the TVS Motor equity research dashboard.

Fetches today's live market data once, then runs the valuation model 27
times — every combination of Operating case x Terminal growth case x WACC
adjustment (3 x 3 x 3) — recalculating each with LibreOffice headless and
extracting every downstream output into data/outputs.json for the Streamlit
dashboard to read.

Every external call (yfinance, web scrape, LibreOffice subprocess) is wrapped
so that one failure never crashes the whole run.
"""

import json
import logging
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger("refresh")

ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = ROOT / "TVS_Motor_Valuation_Model.xlsx"
WORKING_XLSX = ROOT / "model_working.xlsx"
OUTPUTS_JSON = ROOT / "data" / "outputs.json"

# ---------------------------------------------------------------------------
# Cell map — from the Phase 1 openpyxl audit of TVS_Motor_Valuation_Model.xlsx,
# re-verified directly against the workbook's data validations for this change.
# ---------------------------------------------------------------------------

# Market inputs written daily
CELL_CMP = ("Debt & Cost of Capital", "C48")          # TVSMOTOR.NS current market price
CELL_RF = ("Guidance & Assumptions", "C20")            # India 10-year G-Sec yield
PEER_CELLS = {
    "bajaj_auto": {"sheet": "Peer Comps", "cell": "H8", "ticker": "BAJAJ-AUTO.NS", "name": "Bajaj Auto"},
    "hero_motocorp": {"sheet": "Peer Comps", "cell": "H9", "ticker": "HEROMOTOCO.NS", "name": "Hero MotoCorp"},
    "eicher_motors": {"sheet": "Peer Comps", "cell": "H10", "ticker": "EICHERMOT.NS", "name": "Eicher Motors"},
}
TVS_TICKER = "TVSMOTOR.NS"
RF_TICKER = "IN10Y.NS"

# Three live-driven scenario switches — all on Control Panel, verified via
# openpyxl against ws.data_validations.dataValidation:
#   C7  Operating case          -> Conservative / Base / Aggressive
#   C8  Terminal growth case    -> Conservative / Base / Aggressive
#   C10 WACC adjustment         -> -100 bp / -50 bp / As calculated / +50 bp / +100 bp
# The dashboard uses Bull/Base/Bear (Operating & Terminal growth) and
# High/Base/Low (WACC) as friendlier equity-research labels for the same
# underlying Excel dropdown values, so these dicts map dashboard label ->
# literal Excel string that must be written to satisfy each cell's list
# data validation. WACC adjustment intentionally uses only the +/-50bp
# points, not the +/-100bp extremes.
OPERATING_CASE_CELL = ("Control Panel", "C7")
TERMINAL_GROWTH_CELL = ("Control Panel", "C8")
WACC_ADJUSTMENT_CELL = ("Control Panel", "C10")

CASE_LABEL_TO_EXCEL_VALUE = {
    "Bull": "Aggressive",
    "Base": "Base",
    "Bear": "Conservative",
}
WACC_LABEL_TO_EXCEL_VALUE = {
    "High": "+50 bp",
    "Base": "As calculated",
    "Low": "-50 bp",
}
WACC_LABEL_DISPLAY = {
    "High": "+50bps",
    "Base": "As calculated",
    "Low": "-50bps",
}
LABEL_ORDER = ["Bull", "Base", "Bear"]
WACC_LABEL_ORDER = ["High", "Base", "Low"]
ACTIVE_SCENARIO_DEFAULT = {"operating_case": "Base", "terminal_growth_case": "Base", "wacc_adjustment": "Base"}

# Key outputs — all on Control Panel, "B. LIVE OUTPUT" block
OUT_KE = ("Control Panel", "C18")
OUT_WACC = ("Control Panel", "C19")
OUT_TERMINAL_G = ("Control Panel", "C20")
OUT_CORE_AUTO_PS = ("Control Panel", "C21")
OUT_TVS_CREDIT_PS = ("Control Panel", "C22")
OUT_SOTP_PS = ("Control Panel", "C23")
OUT_CONCLUDED_PS = ("Control Panel", "C24")
OUT_CMP = ("Control Panel", "C25")
OUT_UPSIDE = ("Control Panel", "C26")
OUT_RECOMMENDATION = ("Control Panel", "C27")

# All 9 scenario switches — Control Panel!C7:C15, each with a list data
# validation. C7/C8/C10 are the three live-driven ones above; the remaining
# six are recorded once (from the Bull=Base/Base/Base run) as metadata only —
# the dashboard no longer renders them.
SCENARIO_SWITCHES = {
    "Operating case": "C7",
    "Terminal growth case": "C8",
    "Beta basis": "C9",
    "WACC adjustment": "C10",
    "Commodity pass-through": "C11",
    "TVS Credit valuation method": "C12",
    "Peer multiple basis": "C13",
    "SOTP weighting scheme": "C14",
    "Monte Carlo volatility regime": "C15",
}
SCENARIO_SHEET = "Control Panel"
FALLBACK_OPTIONS = {
    "Operating case": ["Conservative", "Base", "Aggressive"],
    "Terminal growth case": ["Conservative", "Base", "Aggressive"],
    "Beta basis": ["Regression beta", "Bottom-up relevered", "Average of both"],
    "WACC adjustment": ["-100 bp", "-50 bp", "As calculated", "+50 bp", "+100 bp"],
    "Commodity pass-through": ["Weak - 40%", "Base - 70%", "Strong - 90%"],
    "TVS Credit valuation method": ["Average of three", "Peer P/B only", "Warranted P/B only", "Transaction only"],
    "Peer multiple basis": ["Peer median", "Peer mean", "Peer low", "Peer high"],
    "SOTP weighting scheme": ["DCF-led", "Balanced", "Market-led"],
    "Monte Carlo volatility regime": ["Low", "Base", "High"],
}

# SOTP waterfall legs — SOTP & Football Field
SOTP_SHEET = "SOTP & Football Field"
SOTP_CORE_AUTO_EV = "C7"          # core auto enterprise value (₹ cr)
SOTP_NET_DEBT = "C8"              # net debt, already negative (₹ cr)
SOTP_SHARES_OUT = "C23"           # shares outstanding (cr)
SOTP_TVS_CREDIT_PS = "D13"
SOTP_OTHER_SUBS_PS = "D15"
SOTP_TOTAL_PS = "D22"

# Football field — SOTP & Football Field!C29:E35 (low, base/mid, high)
FOOTBALL_FIELD_ROWS = {
    "DCF - FCFF": 29,
    "DCF - FCFE": 30,
    "EV/EBITDA": 31,
    "P/E": 32,
    "Precedent Transactions": 33,
    "Asset-Based NAV": 34,
    "DDM": 35,
}

# Sensitivity grid — Sensitivity & Scenarios: WACC rows C27:C31, g cols D26:H26, grid D27:H31
SENS_SHEET = "Sensitivity & Scenarios"
SENS_WACC_CELLS = ["C27", "C28", "C29", "C30", "C31"]
SENS_G_CELLS = ["D26", "E26", "F26", "G26", "H26"]
SENS_GRID_COLS = ["D", "E", "F", "G", "H"]
SENS_GRID_ROWS = [27, 28, 29, 30, 31]

# Monte Carlo
MC_SHEET = "Monte Carlo"
MC_MEAN = "C17"
MC_MEDIAN = "C18"
MC_STD = "C19"
MC_P5 = "C20"
MC_P25 = "C21"
MC_P75 = "C22"
MC_P95 = "C23"
MC_CMP = "C24"
MC_PROB_ABOVE_CMP = "C25"
MC_TRIAL_COL = "J"
MC_TRIAL_FIRST_ROW = 31
MC_TRIAL_LAST_ROW = 1030

# Peer multiples — Peer Comps!C15:D17 (EV/EBITDA, P/E)
PEER_MULTIPLE_ROWS = {"bajaj_auto": 15, "hero_motocorp": 16, "eicher_motors": 17}


def fetch_price(ticker: str):
    """Best-effort last price fetch via yfinance. Returns float or None."""
    try:
        import yfinance as yf

        t = yf.Ticker(ticker)
        try:
            price = t.fast_info.get("lastPrice") or t.fast_info.get("last_price")
            if price:
                return float(price)
        except Exception:
            pass
        info = t.info
        price = info.get("currentPrice") or info.get("regularMarketPrice")
        if price:
            return float(price)
        hist = t.history(period="5d")
        if not hist.empty:
            return float(hist["Close"].iloc[-1])
    except Exception as e:
        log.warning("Price fetch failed for %s: %s", ticker, e)
    return None


def fetch_market_cap(ticker: str):
    """Best-effort market cap fetch via yfinance, in raw INR. Returns float or None."""
    try:
        import yfinance as yf

        t = yf.Ticker(ticker)
        try:
            mc = t.fast_info.get("marketCap") or t.fast_info.get("market_cap")
            if mc:
                return float(mc)
        except Exception:
            pass
        info = t.info
        mc = info.get("marketCap")
        if mc:
            return float(mc)
    except Exception as e:
        log.warning("Market cap fetch failed for %s: %s", ticker, e)
    return None


# Hardcoded floor for the India 10-year G-Sec yield, used only if every live
# source fails (yfinance IN10Y.NS, investing.com, FBIL REST). Approximate
# yield as of August 2026 — fallback, update manually if yield shifts significantly.
RF_HARDCODED_FALLBACK = 0.0685

INVESTING_COM_URL = "https://www.investing.com/rates-bonds/india-10-year-bond-yield"
INVESTING_COM_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.google.com/",
}

# Undocumented — FBIL's public site is a JS-rendered Angular SPA with no known
# stable public REST path, so this is a best-effort probe of plausible endpoints.
FBIL_REST_CANDIDATES = [
    "https://www.fbil.org.in/rest/OverNightRateHistory",
    "https://www.fbil.org.in/api/RatesData",
]


def _fetch_rate_from_investing_com():
    """Best-effort scrape of investing.com's India 10Y bond yield page."""
    import re

    import requests
    from bs4 import BeautifulSoup

    resp = requests.get(INVESTING_COM_URL, headers=INVESTING_COM_HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    tag = soup.find(attrs={"data-test": "instrument-price-last"})
    if tag and tag.text.strip():
        return float(tag.text.strip().replace(",", "")) / 100

    tag = soup.find(id="last_last")
    if tag and tag.text.strip():
        return float(tag.text.strip().replace(",", "")) / 100

    match = re.search(r'"last"\s*:\s*"?(\d{1,2}\.\d{2,4})"?', resp.text)
    if match:
        return float(match.group(1)) / 100

    match = re.search(r"India\s*10[- ]Year.{0,80}?(\d{1,2}\.\d{2,4})\s*%", soup.get_text(" "))
    if match:
        return float(match.group(1)) / 100

    return None


def _fetch_rate_from_fbil_rest():
    """Best-effort probe of undocumented FBIL REST endpoints for the benchmark yield."""
    import requests

    for url in FBIL_REST_CANDIDATES:
        try:
            resp = requests.get(url, headers={"Accept": "application/json"}, timeout=10)
            if resp.status_code != 200:
                continue
            payload = resp.json()
            # Shape is unknown/undocumented — probe common key names defensively.
            candidates = payload if isinstance(payload, list) else [payload]
            for item in candidates:
                if not isinstance(item, dict):
                    continue
                for key in ("rate", "value", "yield", "Rate", "Value"):
                    if key in item:
                        try:
                            val = float(item[key])
                            return val / 100 if val > 1 else val
                        except (TypeError, ValueError):
                            continue
        except Exception:
            continue
    return None


def fetch_risk_free_rate():
    """
    India 10-year G-Sec yield. Source order: yfinance IN10Y.NS, then
    investing.com, then FBIL's (undocumented) REST API, then a hardcoded
    fallback. Always returns a decimal (e.g. 0.0678) — never None — so the
    workbook always gets a value even if every live source is unreachable.
    """
    try:
        import yfinance as yf

        t = yf.Ticker(RF_TICKER)
        hist = t.history(period="5d")
        if not hist.empty:
            val = float(hist["Close"].iloc[-1])
            # Yahoo yield tickers are usually already in percent terms (e.g. 6.78)
            rate = val / 100 if val > 1 else val
            log.info("Risk-free rate from yfinance %s: %.4f", RF_TICKER, rate)
            return rate
    except Exception as e:
        log.warning("yfinance risk-free fetch failed (%s): %s", RF_TICKER, e)

    try:
        rate = _fetch_rate_from_investing_com()
        if rate:
            log.info("Risk-free rate from investing.com: %.4f", rate)
            return rate
        log.warning("investing.com fetched but no rate pattern found")
    except Exception as e:
        log.warning("investing.com scrape failed (likely blocked by Cloudflare bot protection): %s", e)

    try:
        rate = _fetch_rate_from_fbil_rest()
        if rate:
            log.info("Risk-free rate from FBIL REST: %.4f", rate)
            return rate
        log.warning("FBIL REST probe returned no usable data (endpoints are undocumented/unstable)")
    except Exception as e:
        log.warning("FBIL REST probe failed: %s", e)

    log.warning(
        "All live risk-free rate sources failed — using hardcoded fallback %.2f%% "
        "(update RF_HARDCODED_FALLBACK manually if the yield shifts significantly)",
        RF_HARDCODED_FALLBACK * 100,
    )
    return RF_HARDCODED_FALLBACK


def write_market_inputs(wb, results):
    """Write fetched market data into the workbook. wb is loaded with data_only=False."""
    sheet, cell = CELL_CMP
    if results["cmp"] is not None:
        wb[sheet][cell] = results["cmp"]
    else:
        log.warning("CMP unavailable — leaving %s!%s unchanged", sheet, cell)

    sheet, cell = CELL_RF
    if results["rf_rate"] is not None:
        wb[sheet][cell] = results["rf_rate"]
    else:
        log.warning("Rf rate unavailable — leaving %s!%s unchanged", sheet, cell)

    for key, meta in PEER_CELLS.items():
        mc_cr = results["peers"][key]["market_cap_cr"]
        if mc_cr is not None:
            wb[meta["sheet"]][meta["cell"]] = mc_cr
        else:
            log.warning("%s market cap unavailable — leaving %s!%s unchanged", meta["name"], meta["sheet"], meta["cell"])


def write_operating_case(wb, excel_value):
    sheet, cell = OPERATING_CASE_CELL
    wb[sheet][cell] = excel_value


def write_terminal_growth_case(wb, excel_value):
    sheet, cell = TERMINAL_GROWTH_CELL
    wb[sheet][cell] = excel_value


def write_wacc_adjustment(wb, excel_value):
    sheet, cell = WACC_ADJUSTMENT_CELL
    wb[sheet][cell] = excel_value


# LibreOffice's default policy for recalculating formulas loaded from a
# "foreign" format (xlsx) is "prompt the user" — and headless mode can't show
# a prompt, so by default it silently SKIPS recalculation and just passes
# through whatever cached values were already in the file (which, since
# openpyxl wipes cached values on save, means every formula cell comes back
# blank). Forcing OOXMLRecalcMode/ODFRecalcMode to 2 ("Always") in a fresh,
# isolated user profile makes it actually recalculate before saving.
RECALC_ALWAYS_XCU = """<?xml version="1.0" encoding="UTF-8"?>
<oor:items xmlns:oor="http://openoffice.org/2001/registry" xmlns:xs="http://www.w3.org/2001/XMLSchema" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="OOXMLRecalcMode" oor:op="fuse"><value>2</value></prop></item>
 <item oor:path="/org.openoffice.Office.Calc/Formula/Load"><prop oor:name="ODFRecalcMode" oor:op="fuse"><value>2</value></prop></item>
</oor:items>
"""


def _make_libreoffice_profile():
    """A fresh, isolated LibreOffice user profile with recalc-always preset (see RECALC_ALWAYS_XCU)."""
    profile_dir = Path(tempfile.mkdtemp(prefix="lo_profile_"))
    user_dir = profile_dir / "user"
    user_dir.mkdir(parents=True, exist_ok=True)
    (user_dir / "registrymodifications.xcu").write_text(RECALC_ALWAYS_XCU, encoding="utf-8")
    return profile_dir


def recalculate_with_libreoffice():
    """
    Recalculate WORKING_XLSX via LibreOffice headless.

    --convert-to refuses to write when the output path is identical to the
    input path — it fails the internal save (Sfx "Write Code:12") but still
    exits 0, so the file silently stays byte-for-byte unchanged while looking
    like a success. Converting into a separate temp directory and moving the
    result back to WORKING_XLSX avoids that in-place collision.
    """
    soffice = shutil.which("soffice") or shutil.which("soffice.exe")
    if not soffice:
        return False

    profile_dir = _make_libreoffice_profile()
    out_dir = Path(tempfile.mkdtemp(prefix="lo_out_"))
    try:
        profile_uri = profile_dir.resolve().as_uri()
        result = subprocess.run(
            [
                soffice,
                "--headless",
                "--norestore",
                f"-env:UserInstallation={profile_uri}",
                "--calc",
                "--infilter=Calc MS Excel 2007 XML",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(out_dir),
                str(WORKING_XLSX),
            ],
            capture_output=True,
            text=True,
            timeout=180,
        )
        if result.returncode != 0:
            log.warning("LibreOffice recalculation returned code %s: %s", result.returncode, result.stderr[:500])
            return False

        converted = out_dir / WORKING_XLSX.name
        if not converted.exists():
            log.warning("LibreOffice reported success but no output file appeared at %s", converted)
            return False

        shutil.move(str(converted), str(WORKING_XLSX))
        return True
    except Exception as e:
        log.warning("LibreOffice recalculation failed: %s", e)
        return False
    finally:
        shutil.rmtree(profile_dir, ignore_errors=True)
        shutil.rmtree(out_dir, ignore_errors=True)


def safe_get(ws, cell, default=None):
    try:
        val = ws[cell].value
        return val if val is not None else default
    except Exception:
        return default


def make_cell_getter(working_path, fallback_path):
    """
    Open working_path (data_only=True) and fallback_path (data_only=True),
    returning (workbook, get_cell, fallback_used) where get_cell(sheet, cell)
    reads from working_path first and falls back to fallback_path's cached
    value whenever the working copy comes back None.
    """
    try:
        wbv = openpyxl.load_workbook(working_path, data_only=True)
    except Exception as e:
        log.error("Could not open %s for extraction: %s", working_path, e)
        return None, None, []

    wbv_fallback = None
    try:
        wbv_fallback = openpyxl.load_workbook(fallback_path, data_only=True)
    except Exception as e:
        log.warning("Could not open %s for cached-value fallback: %s", fallback_path, e)

    fallback_used = []

    def get_cell(sheet_name, cell):
        val = None
        try:
            if sheet_name in wbv.sheetnames:
                val = safe_get(wbv[sheet_name], cell)
        except Exception:
            val = None
        if val is None and wbv_fallback is not None:
            try:
                if sheet_name in wbv_fallback.sheetnames:
                    fallback_val = safe_get(wbv_fallback[sheet_name], cell)
                    if fallback_val is not None:
                        val = fallback_val
                        fallback_used.append(f"{sheet_name}!{cell}")
            except Exception:
                pass
        return val

    return wbv, get_cell, fallback_used


def extract_scenario_outputs(get_cell):
    """
    Pull the per-combination output block (concluded value, WACC, SOTP,
    football field, sensitivity grid, Monte Carlo, ...) using an
    already-bound get_cell closure from make_cell_getter(). Returns a dict
    shaped for outputs.json["scenarios"][op][tg][wacc].
    """
    ke = get_cell(*OUT_KE)
    wacc = get_cell(*OUT_WACC)
    terminal_g = get_cell(*OUT_TERMINAL_G)
    cost_of_capital = {"ke": ke, "wacc": wacc, "terminal_g": terminal_g}

    core_auto_ps = get_cell(*OUT_CORE_AUTO_PS)
    tvs_credit_ps = get_cell(*OUT_TVS_CREDIT_PS)
    sotp_ps = get_cell(*OUT_SOTP_PS)

    shares_out = get_cell(SOTP_SHEET, SOTP_SHARES_OUT)
    core_auto_ev = get_cell(SOTP_SHEET, SOTP_CORE_AUTO_EV)
    net_debt = get_cell(SOTP_SHEET, SOTP_NET_DEBT)
    other_subs_ps = get_cell(SOTP_SHEET, SOTP_OTHER_SUBS_PS)

    core_auto_ev_ps = (core_auto_ev / shares_out) if (core_auto_ev is not None and shares_out) else None
    net_debt_ps = (net_debt / shares_out) if (net_debt is not None and shares_out) else None

    sotp_breakdown = {
        "core_auto_per_share": core_auto_ps,
        "tvs_credit_per_share": tvs_credit_ps,
        "sotp_per_share": sotp_ps,
        "other_subs_per_share": other_subs_ps,
        "net_debt_per_share": net_debt_ps,
    }

    concluded_value_per_share = get_cell(*OUT_CONCLUDED_PS)
    upside_pct = get_cell(*OUT_UPSIDE)
    recommendation = get_cell(*OUT_RECOMMENDATION)

    # Sensitivity grid
    wacc_values = [get_cell(SENS_SHEET, c) for c in SENS_WACC_CELLS]
    g_values = [get_cell(SENS_SHEET, c) for c in SENS_G_CELLS]
    matrix = []
    for r in SENS_GRID_ROWS:
        row_vals = [get_cell(SENS_SHEET, f"{col}{r}") for col in SENS_GRID_COLS]
        matrix.append(row_vals)
    sensitivity_grid = {"wacc_values": wacc_values, "g_values": g_values, "matrix": matrix}

    # Monte Carlo — percentiles only; raw trial arrays are dropped from the
    # persisted JSON to keep it a sane size across 27 combinations (27,000
    # numbers otherwise). The dashboard reconstructs a normal distribution
    # from mean/std_dev for the histogram when raw_trials is absent.
    mean = get_cell(MC_SHEET, MC_MEAN)
    std_dev = get_cell(MC_SHEET, MC_STD)
    median = get_cell(MC_SHEET, MC_MEDIAN)
    prob_above_cmp = get_cell(MC_SHEET, MC_PROB_ABOVE_CMP)

    trials = []
    try:
        for r in range(MC_TRIAL_FIRST_ROW, MC_TRIAL_LAST_ROW + 1):
            v = get_cell(MC_SHEET, f"{MC_TRIAL_COL}{r}")
            if v is not None:
                trials.append(v)
    except Exception as e:
        log.warning("Could not read Monte Carlo trial data: %s", e)

    if trials:
        try:
            import numpy as np

            p10 = float(np.percentile(trials, 10))
            p25 = float(np.percentile(trials, 25))
            p50 = float(np.percentile(trials, 50))
            p75 = float(np.percentile(trials, 75))
            p90 = float(np.percentile(trials, 90))
        except Exception:
            p10 = p25 = p50 = p75 = p90 = None
    else:
        p10 = None
        p25 = get_cell(MC_SHEET, MC_P25)
        p50 = median
        p75 = get_cell(MC_SHEET, MC_P75)
        p90 = None

    monte_carlo = {
        "p10": p10,
        "p25": p25,
        "p50": p50 if p50 is not None else median,
        "p75": p75,
        "p90": p90,
        "mean": mean,
        "std_dev": std_dev,
        "prob_above_cmp": prob_above_cmp,
        "raw_trials": None,
    }

    # SOTP legs — ordered waterfall
    sotp_legs = [
        {"name": "Core Automotive (EV)", "value_per_share": core_auto_ev_ps, "type": "add"},
        {"name": "Net Debt", "value_per_share": net_debt_ps, "type": "subtract"},
        {"name": "TVS Credit (85.15%)", "value_per_share": tvs_credit_ps, "type": "add"},
        {"name": "Other Subsidiaries", "value_per_share": other_subs_ps, "type": "add"},
    ]

    # Football field
    football_field = []
    for methodology, row in FOOTBALL_FIELD_ROWS.items():
        low = get_cell(SOTP_SHEET, f"C{row}")
        mid = get_cell(SOTP_SHEET, f"D{row}")
        high = get_cell(SOTP_SHEET, f"E{row}")
        football_field.append({"methodology": methodology, "low": low, "mid": mid, "high": high})

    return {
        "concluded_value_per_share": concluded_value_per_share,
        "wacc": wacc,
        "ke": ke,
        "terminal_g": terminal_g,
        "upside_pct": upside_pct,
        "recommendation": recommendation,
        "sotp_breakdown": sotp_breakdown,
        "sotp_legs": sotp_legs,
        "football_field": football_field,
        "sensitivity_grid": sensitivity_grid,
        "monte_carlo": monte_carlo,
        "cost_of_capital": cost_of_capital,
    }


def extract_market_inputs(get_cell, results):
    """Market inputs are shared across all 27 combinations (peer multiples don't move with them)."""
    market_inputs = {
        "cmp": results.get("cmp"),
        "rf_rate": results.get("rf_rate"),
        "peers": {
            key: {
                "ticker": meta["ticker"],
                "name": meta["name"],
                "price": results["peers"][key].get("price"),
                "currency": "INR",
            }
            for key, meta in PEER_CELLS.items()
        },
    }
    for key, row in PEER_MULTIPLE_ROWS.items():
        market_inputs["peers"][key]["ev_ebitda"] = get_cell("Peer Comps", f"C{row}")
        market_inputs["peers"][key]["pe"] = get_cell("Peer Comps", f"D{row}")
    return market_inputs


def extract_scenario_switches(working_path, get_cell):
    """
    All 9 Control Panel switches with their valid dropdown options, read once
    from the Bull=Base/Base/Base run — metadata only. The dashboard drives
    Operating case, Terminal growth case and WACC adjustment itself via the
    scenarios[op][tg][wacc] structure, not this block.
    """
    dv_options = {}
    try:
        wbf = openpyxl.load_workbook(working_path, data_only=False)
        ws_formula = wbf[SCENARIO_SHEET]
        for dv in ws_formula.data_validations.dataValidation:
            if dv.type == "list" and dv.formula1:
                opts = dv.formula1.strip('"').split(",")
                for coord in str(dv.sqref).split():
                    dv_options[coord] = opts
    except Exception as e:
        log.warning("Could not read data validations: %s", e)

    scenario_switches = {}
    for name, cell in SCENARIO_SWITCHES.items():
        current_value = get_cell(SCENARIO_SHEET, cell)
        options = dv_options.get(cell) or FALLBACK_OPTIONS.get(name, [])
        scenario_switches[name] = {"current_value": current_value, "valid_options": options}
    return scenario_switches


def run_scenario(op_label, tg_label, wacc_label, results, run_index, total_runs):
    """
    Build a fresh working copy FROM THE CLEAN ORIGINAL (never chained on a
    previous run's copy), write market inputs plus this combination's three
    switch values, recalculate, and extract its output block.
    Returns (scenario_outputs, get_cell, recalculated, fallback_used).
    """
    op_excel = CASE_LABEL_TO_EXCEL_VALUE[op_label]
    tg_excel = CASE_LABEL_TO_EXCEL_VALUE[tg_label]
    wacc_excel = WACC_LABEL_TO_EXCEL_VALUE[wacc_label]

    log.info(
        "Running [%d/%d]: Operating=%s | Terminal g=%s | WACC=%s (%s)",
        run_index, total_runs, op_label, tg_label, wacc_label, WACC_LABEL_DISPLAY[wacc_label],
    )

    shutil.copy(SOURCE_XLSX, WORKING_XLSX)

    try:
        wb = openpyxl.load_workbook(WORKING_XLSX, keep_vba=False, data_only=False)
        write_market_inputs(wb, results)
        write_operating_case(wb, op_excel)
        write_terminal_growth_case(wb, tg_excel)
        write_wacc_adjustment(wb, wacc_excel)
        wb.save(WORKING_XLSX)
    except Exception as e:
        log.error(
            "  [%s/%s/%s] Failed to write inputs into workbook: %s", op_label, tg_label, wacc_label, e
        )

    try:
        recalculated = recalculate_with_libreoffice()
    except Exception as e:
        log.warning("  LibreOffice call raised unexpectedly: %s", e)
        recalculated = False
    log.info("  LibreOffice recalc: %s", "OK" if recalculated else "SKIPPED (fallback)")

    try:
        wbv, get_cell, fallback_used = make_cell_getter(WORKING_XLSX, SOURCE_XLSX)
    except Exception as e:
        log.error("  [%s/%s/%s] Extraction setup failed: %s", op_label, tg_label, wacc_label, e)
        return None, None, recalculated, []

    if get_cell is None:
        return None, None, recalculated, []

    try:
        scenario_outputs = extract_scenario_outputs(get_cell)
    except Exception as e:
        log.error("  [%s/%s/%s] Output extraction failed: %s", op_label, tg_label, wacc_label, e)
        return None, None, recalculated, fallback_used

    concluded = scenario_outputs.get("concluded_value_per_share")
    rec = scenario_outputs.get("recommendation")
    concluded_str = f"₹{concluded:,.2f}" if isinstance(concluded, (int, float)) else "N/A"
    log.info("  Concluded value: %s | Recommendation: %s", concluded_str, rec or "N/A")

    if fallback_used:
        non_trial = [c for c in fallback_used if not c.startswith(f"{MC_SHEET}!{MC_TRIAL_COL}")]
        if non_trial:
            log.warning(
                "  %d non-trial cell(s) fell back to the source workbook's cached values: %s",
                len(non_trial), non_trial[:15],
            )

    return scenario_outputs, get_cell, recalculated, fallback_used


def main():
    log.info("=== TVS Motor dashboard daily refresh starting ===")

    if not SOURCE_XLSX.exists():
        log.error("Source workbook not found at %s — cannot proceed", SOURCE_XLSX)
        sys.exit(1)

    results = {"cmp": None, "rf_rate": None, "peers": {}}
    fetch_status = {}

    results["cmp"] = fetch_price(TVS_TICKER)
    fetch_status["TVSMOTOR CMP"] = results["cmp"] is not None

    for key, meta in PEER_CELLS.items():
        price = fetch_price(meta["ticker"])
        mc = fetch_market_cap(meta["ticker"])
        mc_cr = (mc / 1e7) if mc is not None else None
        results["peers"][key] = {"price": price, "market_cap_cr": mc_cr}
        fetch_status[f"{meta['name']} price/mcap"] = price is not None and mc_cr is not None

    results["rf_rate"] = fetch_risk_free_rate()
    fetch_status["India 10Y G-Sec (Rf)"] = results["rf_rate"] is not None

    # Bull/Base/Bear x Bull/Base/Bear x High/Base/Low = 27 combinations.
    # Base/Base/Base runs first so market_inputs/scenario_switches capture it.
    combos = [(op, tg, wacc) for op in LABEL_ORDER for tg in LABEL_ORDER for wacc in WACC_LABEL_ORDER]
    combos.sort(key=lambda c: (c != ("Base", "Base", "Base"),))
    total_runs = len(combos)

    scenarios = {}
    run_log = []  # (op, tg, wacc, concluded, recommendation, recalculated)
    market_inputs = None
    scenario_switches = None

    for i, (op_label, tg_label, wacc_label) in enumerate(combos, start=1):
        scenario_outputs, get_cell, recalculated, fallback_used = run_scenario(
            op_label, tg_label, wacc_label, results, i, total_runs
        )

        if scenario_outputs is None:
            log.error(
                "  [%s/%s/%s] Extraction failed — combination will be missing from outputs.json",
                op_label, tg_label, wacc_label,
            )
            run_log.append((op_label, tg_label, wacc_label, None, None, recalculated))
            continue

        scenarios.setdefault(op_label, {}).setdefault(tg_label, {})[wacc_label] = scenario_outputs
        run_log.append(
            (
                op_label, tg_label, wacc_label,
                scenario_outputs.get("concluded_value_per_share"),
                scenario_outputs.get("recommendation"),
                recalculated,
            )
        )

        if market_inputs is None and (op_label, tg_label, wacc_label) == ("Base", "Base", "Base"):
            try:
                market_inputs = extract_market_inputs(get_cell, results)
                scenario_switches = extract_scenario_switches(WORKING_XLSX, get_cell)
            except Exception as e:
                log.warning("Could not extract market_inputs/scenario_switches from Base/Base/Base run: %s", e)

    if market_inputs is None:
        # Base/Base/Base run failed entirely — fall back to raw fetch results
        # so the dashboard still has prices even without a valuation to show.
        market_inputs = {
            "cmp": results.get("cmp"),
            "rf_rate": results.get("rf_rate"),
            "peers": {
                key: {"ticker": meta["ticker"], "name": meta["name"], "price": results["peers"][key].get("price"), "currency": "INR"}
                for key, meta in PEER_CELLS.items()
            },
        }
    if scenario_switches is None:
        scenario_switches = {}

    outputs = {
        "valuation_date": datetime.now().strftime("%Y-%m-%d"),
        "last_refreshed": datetime.now(timezone.utc).isoformat(),
        "market_inputs": market_inputs,
        "scenario_switches": scenario_switches,
        "active_scenario": ACTIVE_SCENARIO_DEFAULT,
        "scenarios": scenarios,
    }

    OUTPUTS_JSON.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUTS_JSON, "w", encoding="utf-8") as f:
        json.dump(outputs, f, indent=2, default=str)
    log.info("Wrote %s", OUTPUTS_JSON)

    log.info("=== RUN SUMMARY ===")
    for label, ok in fetch_status.items():
        log.info("  %-30s %s", label, "OK" if ok else "FAILED (kept existing value)")

    log.info("--- All %d combinations ---", total_runs)
    log.info("  %-6s %-6s %-6s %-14s %-10s %s", "Op", "TermG", "WACC", "Concluded (Rs)", "Rec", "Recalc")
    for op_label, tg_label, wacc_label, concluded, rec, recalculated in sorted(
        run_log, key=lambda r: (LABEL_ORDER.index(r[0]), LABEL_ORDER.index(r[1]), WACC_LABEL_ORDER.index(r[2]))
    ):
        concluded_str = f"{concluded:,.2f}" if isinstance(concluded, (int, float)) else "N/A"
        log.info(
            "  %-6s %-6s %-6s %-14s %-10s %s",
            op_label, tg_label, wacc_label, concluded_str, rec or "N/A",
            "OK" if recalculated else "FALLBACK",
        )

    succeeded = sum(len(scenarios.get(op, {}).get(tg, {})) for op in LABEL_ORDER for tg in LABEL_ORDER)
    distinct_values = len(
        {
            scenarios[op][tg][wacc].get("concluded_value_per_share")
            for op in scenarios
            for tg in scenarios[op]
            for wacc in scenarios[op][tg]
            if scenarios[op][tg][wacc].get("concluded_value_per_share") is not None
        }
    )
    log.info(
        "=== %d/%d combinations populated, %d distinct concluded-value results ===",
        succeeded, total_runs, distinct_values,
    )
    log.info("=== Refresh complete ===")


if __name__ == "__main__":
    main()
